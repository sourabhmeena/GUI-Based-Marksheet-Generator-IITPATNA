from logging import debug
from flask import Flask,render_template,request,session,flash,redirect,send_file, make_response
from flask.scaffold import _matching_loader_thinks_module_is_package
# from werkzeug import secure_filename
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import Workbook
import openpyxl
import csv
import os
import shutil
os.system('cls')

app = Flask(__name__)
app.secret_key='super-secret-key'
#--------------------------------------------------------------
@app.route("/")
def hello():
    return render_template("index.html")


#---------------------------------------------------------------
#------------------------------------------------------------------------------------
def generate_concise_marksheet(pmarks,nmarks,no_que,info,stud):
    dict={}
    
    for roll_no in info:
        no_right_ans,no_wrong_ans,not_attempted=0,0,0
        for i in range(no_que):
            if(info['ANSWER'][i]==info[roll_no][i]):
                no_right_ans+=1
            elif(info[roll_no][i]==''):
                not_attempted+=1
            else :
                no_wrong_ans+=1
        dict.update({roll_no:[ str((no_right_ans*pmarks)+(no_wrong_ans*nmarks))+'/'+str(pmarks*no_que),str([no_right_ans,no_wrong_ans,not_attempted])]})
    head=[] 
    list_of_list=[]
    with open(r"responses.csv",'r') as file_response:
        head=['Timestamp','Email address','Google_Score','name','IITP webmail','Phone(10 digit only)','Score_After_Negative','Roll Number'] +['Unnamed: %d'%i for i in range(7,no_que+7)]+['statusAns',]
        h=file_response.readline()
        data=csv.reader(file_response)
        for row in data:
            list_of_list.append(row)
    
    for i in range(len(list_of_list)):
        list_of_list[i].insert(head.index("Score_After_Negative"),dict[list_of_list[i][6]][0])
        list_of_list[i].insert(head.index("statusAns"),dict[list_of_list[i][7]][1])         #7 because now score_after_negative add ho chuka h
    list_marksheet_of_absent_stud=[]
    for key in stud:
        if(info.get(key)==None):
            d=['Absent:','Absent','Absent',stud[key],'Absent','Absent','Absent',key]+['Absent 'for i in range(7,no_que+7)]+['Absent',]
            list_of_list.append(d)
    
    wb=Workbook()
    sheet=wb.active
    sheet.append(head)
    for row in list_of_list:
        sheet.append(row)
    wb.save(r"marksheet\concise_marksheet.xlsx") 

#------------------------------------------------------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------------------------------------------------
def generate_rollno_wise_marksheet(roll_no, pmarks, nmarks,no_que,info,stud):
    no_right_ans,no_wrong_ans,not_attempted=0,0,0
    file_name = roll_no[0:4]+roll_no[4:6].upper()+roll_no[6:]
    for i in range(no_que):
        if(info['ANSWER'][i]==info[roll_no][i]):
            no_right_ans+=1
        elif(info[roll_no][i]==''):
            not_attempted+=1
        else :
            no_wrong_ans+=1
    wb = Workbook()
    sheet = wb.active
    for i in 'ABCDEFG':
        sheet.column_dimensions[i].width = 20
    img = openpyxl.drawing.image.Image('logo.jpg')
    img.anchor = 'A1'
    sheet.add_image(img)
    # -----line 5
    sheet.merge_cells('A5:E5')
    cell = sheet.cell(row=5, column=1)
    cell.value = 'Mark Sheet'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(name='Century', bold=True, size=20, underline='single')
   # ----- line 6
    col_style = Font(name="Century", size=12)
    sheet.cell(6, 1).value = 'Name:'
    sheet.cell(6, 1).font = col_style
    sheet.cell(6, 1).alignment = Alignment(horizontal='right')
    sheet.merge_cells('B6:C6')
    sheet.cell(6, 2).value = stud[roll_no]
    sheet.cell(6, 2).alignment = Alignment(horizontal='left')
    sheet.cell(6, 2).font=Font(name="Century",sz=12, bold=True)
    sheet.cell(6, 4).value = "Exam:"
    sheet.cell(6, 4).font = col_style
    sheet.cell(6, 4).alignment = Alignment(horizontal='right')
    sheet.cell(6, 5).value = 'quiz'
    sheet.cell(6, 5).alignment = Alignment(horizontal='left')
    sheet.cell(6, 5).font=Font(name="Century",sz=12, bold=True)
   #----- line7
    sheet.cell(7, 1).value = 'Roll Number:'
    sheet.cell(7, 1).font = col_style
    sheet.cell(7, 1).alignment = Alignment(horizontal='right')
    sheet.cell(7, 2).value = file_name
    sheet.cell(7, 2).alignment = Alignment(horizontal='left')
    sheet.cell(7, 2).font=Font(name="Century",sz=12, bold=True)
    # ----table
    data = [['', "Right", 'Wrong', "Not Attempt", 'Max'],
            ['No.',(no_right_ans),(no_wrong_ans),(not_attempted),(no_que)], ['Marking',(pmarks),(nmarks),0, ''], ['Total', (no_right_ans*pmarks),(no_wrong_ans*nmarks), '', str((no_right_ans*pmarks)+(no_wrong_ans*nmarks))+'/'+str(pmarks*no_que)]]
    cell_border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000')
                         )
    n_rows=len(data)
    for row in range(9,n_rows+9):
        n_cols=len(data[row-9])
        for col in range(1,n_cols+1):
            cell=sheet.cell(row,col)
            cell.value=data[row-9][col-1]
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border=cell_border
            if(row==9 or col==1):
                cell.font=Font(name="Century",sz=12, bold=True)
            elif(col==2):
                cell.font=Font(name="Century",color="159207",sz=12)
            elif(col==3):
                cell.font=Font(name="Century",color="FD0606",sz=12)
            elif(row==12 and col==5):
                cell.font=Font(name="Century",sz=12,color='0028EF')
            else:
                cell.font=col_style
    #----here start ans table;
     #--table heading---------
    sheet.cell(15,1).value='Student Ans'
    sheet.cell(15,1).font=Font(name="Century",sz=12, bold=True)
    sheet.cell(15,1).alignment = Alignment(horizontal='center',vertical='center')
    sheet.cell(15,1).border=cell_border
    sheet.cell(15,2).value='Correct Ans'
    sheet.cell(15,2).font=Font(name="Century",sz=12, bold=True)
    sheet.cell(15,2).alignment = Alignment(horizontal='center',vertical='center')
    sheet.cell(15,2).border=cell_border
     #----data
    for i in range(no_que):
        if(info[roll_no][i]==info['ANSWER'][i]):
            sheet.cell(16+i,1).value=info[roll_no][i]
            sheet.cell(16+i,1).font=Font(name="Century",sz=12, color='159207')
            sheet.cell(16+i,1).alignment = Alignment(horizontal='center',vertical='center')
            sheet.cell(16+i,1).border=cell_border
            sheet.cell(16+i,2).value=info['ANSWER'][i]
            sheet.cell(16+i,2).font=Font(name="Century",sz=12,color='0028EF')
            sheet.cell(16+i,2).alignment = Alignment(horizontal='center',vertical='center')
            sheet.cell(16+i,2).border=cell_border
        else:
            sheet.cell(16+i,1).value=info[roll_no][i]
            sheet.cell(16+i,1).font=Font(name="Century",color="FD0606",sz=12)
            sheet.cell(16+i,1).alignment = Alignment(horizontal='center',vertical='center')
            sheet.cell(16+i,1).border=cell_border
            sheet.cell(16+i,2).value=info['ANSWER'][i]
            sheet.cell(16+i,2).font=Font(name="Century",sz=12,color='0028EF')
            sheet.cell(16+i,2).alignment = Alignment(horizontal='center',vertical='center')
            sheet.cell(16+i,2).border=cell_border
    
    # here we need to change address  sample_output\ ..
    wb.save(r"marksheet\%s.xlsx" % file_name)

#----------------------------------------------------------------------------------------------------------------------------------------

@app.route("/uploader",methods=['GET','POST'])
def uploader():
    if(request.method=='POST'):
        try:
            f=request.files['master']
            f.save(f.filename)
            flash("file_uploaded")
        except:
            flash("Please Upload file after Selecting")
        return render_template("index.html")
@app.route("/upload",methods=['GET','POST'])
def upload():
    if(request.method=='POST'):
        try:
            f=request.files['response']
            f.save(f.filename)
            flash("file uploaded")
        except:
            flash("Please Upload file after Selecting")
        return render_template("index.html")     
#--------------------------------------------------------------------------------------------
# -------------------------------------------------------------------------------------------    
@app.route("/roll_wise",methods=['GET','POST'])
def roll_wise():
    if(os.path.isfile(r"responses.csv")==False):
        flash("first upload responses.csv")
        return render_template("index.html")
    if(os.path.isfile(r"master_roll.csv")==False):
        flash("first upload master_roll.csv")
        return render_template("index.html")
    if(request.method=='POST'):
        pmarks=0
        nmarks=0
        try:
            pmarks=int(request.form["pmarks"])
            nmarks=int(request.form["nmarks"])
        except:
            flash("It seems like You forgot to give marks for right or wrong Question")
            return render_template('index.html')
        info ={} 
        stud ={}
        no_que=0
        with open(r"responses.csv",'r') as file_response:              #no of questions
            data=csv.reader(file_response)
            for row in data:
                if(row[6]=='ANSWER'):
                    no_que=len(row)-7
                    break

        with open(r"responses.csv",'r') as file_response:
            head=file_response.readline()
            data=csv.reader(file_response)
            for row in data:
                info.update({row[6]:row[7:no_que+7]})   # 'rollno' : ['obtion A', 'obtion B' ,.............];
        # print(len(info['ANSWER']))
        # exit()
        with open(r"master_roll.csv",'r') as file_master:
            head=file_master.readline()
            data=csv.reader(file_master)
            for row in data:
                stud.update({row[0]:row[1]})     # 'rollno' : 'name' ;
        answer=info.get('ANSWER')
        if(answer==None):
            flash("ANSWER is not present in response.csv file, cann't proceed",'error')
        else:
            for roll_no in stud:
                generate_rollno_wise_marksheet(roll_no,pmarks,nmarks,no_que,info,stud)
            flash(" roll no wise marksheet generated  ",'info')
    return render_template("index.html")

#--------------------------------------------------------------------------------------------
# -------------------------------------------------------------------------------------------    
@app.route("/concise",methods=['GET','POST'])
def concise():
    if(os.path.isfile(r"responses.csv")==False):
        flash("first upload responses.csv")
        return render_template("index.html")
    if(os.path.isfile(r"master_roll.csv")==False):
        flash("first upload master_roll.csv")
        return render_template("index.html")
    if(request.method=='POST'):
        pmarks=0
        nmarks=0
        try:
            pmarks=int(request.form["pmarks"])
            nmarks=int(request.form["nmarks"])
        except:
            flash("It seems like You forgot to give marks for right or wrong Question")
            return render_template('index.html')
        info ={} 
        stud ={}
        no_que=0
        with open(r"responses.csv",'r') as file_response:              #no of questions
            data=csv.reader(file_response)
            for row in data:
                if(row[6]=='ANSWER'):
                    no_que=len(row)-7
                    break

        with open(r"responses.csv",'r') as file_response:
            head=file_response.readline()
            data=csv.reader(file_response)
            for row in data:
                info.update({row[6]:row[7:no_que+7]})   # 'rollno' : ['obtion A', 'obtion B' ,.............];
        # print(len(info['ANSWER']))
        # exit()
        with open(r"master_roll.csv",'r') as file_master:
            head=file_master.readline()
            data=csv.reader(file_master)
            for row in data:
                stud.update({row[0]:row[1]})     # 'rollno' : 'name' ;
        answer=info.get('ANSWER')
        if(answer==None):
            flash("ANSWER is not present in response.csv file, cann't proceed")
        else:
            generate_concise_marksheet(pmarks,nmarks,no_que,info ,stud)
            flash(" concise marksheet generated  ",'info')
    return render_template("index.html")

@app.route("/download",methods=['POST'])
def download():
    if(os.path.exists("./marksheet.zip")):
        os.remove("./marksheet.zip")
    shutil.make_archive("marksheet",'zip',"./marksheet")
    path="./marksheet.zip"
    return send_file(path,as_attachment=True)

if __name__=="__main__":
    app.run(debug=True,port=1000)

